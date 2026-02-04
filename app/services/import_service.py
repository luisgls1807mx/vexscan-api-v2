"""
VexScan API - Import Service v2 (Optimizado)
Usa RPC bulk processing para máximo rendimiento
"""

from typing import Optional, Dict, Any, List
from datetime import datetime
from uuid import uuid4
import hashlib
import logging
import io

from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment

from app.core.supabase import supabase
from app.core.postgres import get_postgres_client
from app.core.config import settings
from app.core.exceptions import (
    ParseError, 
    StorageError, 
    RPCError, 
    DuplicateError,
    ValidationError
)
from app.adapters import AdapterRegistry, get_adapter_for_file, ScanResult

logger = logging.getLogger(__name__)


# Excel severity colors
SEVERITY_COLORS = {
    "Critical": "FF0000",
    "High": "FF6600",
    "Medium": "FFCC00",
    "Low": "00CC00",
    "Info": "0066FF",
}


class ImportService:
    """
    Service for processing and importing scan files.
    
    v2 Changes:
    - Uses bulk RPC function (fn_process_scan_import_v3)
    - Sends ALL adapter fields to database
    - 10-50x faster for large files
    - Supports batch processing for very large files
    """
    
    # Threshold to use batch processing
    BATCH_THRESHOLD = 10   # Más de 5k findings = usar batches (evita timeouts)
    BATCH_SIZE = 1000         # Procesar de 250 en 250 (evita timeouts en BD)
    
    async def process_scan(
        self,
        access_token: str,
        workspace_id: str,
        file_content: bytes,
        filename: str,
        project_id: Optional[str] = None,
        network_zone: str = "internal",
        scanner_hint: Optional[str] = None
    ) -> Dict[str, Any]:
        """
        Complete scan import workflow using optimized RPC.
        
        1. Validate file
        2. Check for duplicates  
        3. Upload to storage
        4. Parse using adapter
        5. Send to RPC for bulk processing (single or batch mode)
        6. Return summary
        """
        # 1. Calculate file hash
        file_hash = hashlib.sha256(file_content).hexdigest()
        file_size = len(file_content)
        
        # 2. Check for duplicate
        is_duplicate = await self._check_duplicate(
            access_token, workspace_id, file_hash, project_id
        )
        if is_duplicate:
            raise DuplicateError("Scan file", filename)
        
        # 3. Detect and validate scanner
        adapter = get_adapter_for_file(filename, file_content, scanner_hint)
        logger.info(f"Using adapter: {adapter.scanner_name}")
        
        is_valid = await adapter.validate(file_content, filename)
        if not is_valid:
            raise ParseError(
                f"Invalid {adapter.scanner_name} file format",
                scanner=adapter.scanner_name,
                filename=filename
            )
        
        # 4. Upload to storage
        storage_path = await self._upload_to_storage(
            workspace_id, file_content, filename
        )
        
        try:
            # 5. Parse file
            logger.info(f"Parsing {filename}...")
            scan_result = await adapter.parse(file_content, filename)
            logger.info(
                f"Parsed {scan_result.total_findings} findings from "
                f"{scan_result.total_hosts} hosts"
            )
            
            # 6. Decidir modo de procesamiento
            if scan_result.total_findings > self.BATCH_THRESHOLD:
                # Modo batch para archivos grandes
                logger.info(f"Using BATCH mode for {scan_result.total_findings} findings")
                return await self._process_in_batches(
                    access_token=access_token,
                    workspace_id=workspace_id,
                    project_id=project_id,
                    filename=filename,
                    storage_path=storage_path,
                    file_size=file_size,
                    file_hash=file_hash,
                    network_zone=network_zone,
                    scan_result=scan_result,
                    adapter=adapter
                )
            else:
                # Modo single para archivos normales
                return await self._process_single(
                    access_token=access_token,
                    workspace_id=workspace_id,
                    project_id=project_id,
                    filename=filename,
                    storage_path=storage_path,
                    file_size=file_size,
                    file_hash=file_hash,
                    network_zone=network_zone,
                    scan_result=scan_result,
                    adapter=adapter
                )
            
        except Exception as e:
            logger.error(f"Import failed: {e}")
            # Try to clean up storage
            try:
                supabase.service.storage.from_(settings.STORAGE_BUCKET).remove([storage_path])
            except:
                pass
            raise
    
    async def _process_single(
        self,
        access_token: str,
        workspace_id: str,
        project_id: Optional[str],
        filename: str,
        storage_path: str,
        file_size: int,
        file_hash: str,
        network_zone: str,
        scan_result: ScanResult,
        adapter
    ) -> Dict[str, Any]:
        """
        Procesar archivo con traducción integrada.
        
        FLUJO OPTIMIZADO:
        1. Serializar assets y findings
        2. Extraer vulnerabilidades únicas del archivo
        3. Consultar BD: ¿cuáles ya existen? (una sola vez)
        4. Traducir SOLO las nuevas (ignorando campos vacíos/N/A)
        5. Insertar nuevas al catálogo
        6. Asignar vulnerability_ids a findings
        7. Procesar con RPC
        """
        from app.services.translation_service import translation_service
        
        # 1. Serializar assets y findings
        assets_json = self._serialize_assets(scan_result.assets)
        findings_json = self._serialize_findings(scan_result.findings)
        
        # 2. CATÁLOGO Y TRADUCCIÓN: Procesar ANTES de enviar a BD
        catalog_stats = {
            'total_unique': 0,
            'already_existed': 0,
            'new_translated': 0,
            'new_without_translation': 0
        }
        
        try:
            # Extraer vulnerabilidades únicas del archivo
            unique_vulns = self._extract_unique_vulnerabilities(findings_json)
            
            if unique_vulns:
                logger.info(f"Processing {len(unique_vulns)} unique vulnerabilities...")
                
                # Proceso completo: consultar existentes, traducir nuevas, insertar
                result = await translation_service.translate_new_vulnerabilities(
                    access_token=access_token,
                    scanner=adapter.scanner_name,
                    vulnerabilities=unique_vulns
                )
                
                plugin_to_vuln_id = result['plugin_to_vuln_id']
                catalog_stats = result['stats']
                
                # Asignar vulnerability_id a cada finding
                findings_json = self._assign_vulnerability_ids(findings_json, plugin_to_vuln_id)
        
        except Exception as e:
            logger.warning(f"Translation step failed (continuing without): {e}")
        
        # 3. Procesar scan con RPC
        import anyio
        result = await anyio.to_thread.run_sync(
            lambda: supabase.rpc_with_token(
                'fn_process_scan_import_v4',
                access_token,
                {
                    'p_workspace_id': workspace_id,
                    'p_project_id': project_id,
                    'p_file_name': filename,
                    'p_storage_path': storage_path,
                    'p_file_size': file_size,
                    'p_file_hash': file_hash,
                    'p_scanner': adapter.scanner_name,
                    'p_network_zone': network_zone,
                    'p_uploaded_by': None,
                    'p_scan_start': scan_result.scan_start.isoformat() if scan_result.scan_start else None,
                    'p_scan_end': scan_result.scan_end.isoformat() if scan_result.scan_end else None,
                    'p_assets': assets_json,
                    'p_findings': findings_json
                }
            )
        )
        
        if not result.get('success'):
            error_msg = result.get('error', 'Unknown error')
            raise RPCError('fn_process_scan_import_v4', error_msg)
        
        summary = result.get('summary', {})
        logger.info(
            f"Import completed in {result.get('processing_time_ms', 0)}ms. "
            f"Findings: {summary.get('findings_created', 0)} new, {summary.get('findings_updated', 0)} updated. "
            f"Catalog: {catalog_stats.get('new_translated', 0)} translated, "
            f"{catalog_stats.get('already_existed', 0)} reused"
        )
        
        return {
            "scan_import_id": result.get('scan_import_id'),
            "scanner": adapter.scanner_name,
            "status": "processed",
            "processing_time_ms": result.get('processing_time_ms'),
            "mode": "single",
            **result.get('summary', {}),
            "catalog_stats": catalog_stats,
            "findings_by_severity": result.get('findings_by_severity', {}),
            "scan_info": {
                "name": scan_result.scan_name,
                "policy": scan_result.scan_policy,
                "start": scan_result.scan_start.isoformat() if scan_result.scan_start else None,
                "end": scan_result.scan_end.isoformat() if scan_result.scan_end else None,
            },
            "warnings": scan_result.warnings,
            "errors": scan_result.errors
        }
    
    def _extract_unique_vulnerabilities(self, findings_json: List[Dict]) -> List[Dict]:
        """Extrae vulnerabilidades únicas de los findings para el catálogo."""
        unique = {}
        for f in findings_json:
            pid = f.get('scanner_finding_id') or f.get('plugin_id')
            if pid and pid not in unique:
                unique[pid] = {
                    'plugin_id': str(pid),
                    'scanner_finding_id': str(pid),
                    'title': f.get('title'),
                    'synopsis': f.get('synopsis'),
                    'description': f.get('description'),
                    'solution': f.get('solution'),
                    'plugin_output': f.get('plugin_output'),  # Agregar plugin_output
                    'severity': f.get('severity'),
                    'cwe': f.get('cwe'),
                    'plugin_family': f.get('plugin_family'),
                    'cvss_score': f.get('cvss_score'),
                    'cvss_vector': f.get('cvss_vector'),
                    'cvss3_score': f.get('cvss3_score'),
                    'cvss3_vector': f.get('cvss3_vector'),
                }
        return list(unique.values())
    
    def _assign_vulnerability_ids(
        self, 
        findings_json: List[Dict], 
        plugin_to_vuln_id: Dict[str, int]
    ) -> List[Dict]:
        """Asigna vulnerability_id a cada finding basado en su plugin_id."""
        for f in findings_json:
            pid = f.get('scanner_finding_id') or f.get('plugin_id')
            if pid:
                vuln_id = plugin_to_vuln_id.get(str(pid))
                if vuln_id:
                    f['vulnerability_id'] = vuln_id
        return findings_json

    
    async def _process_in_batches(
        self,
        access_token: str,
        workspace_id: str,
        project_id: Optional[str],
        filename: str,
        storage_path: str,
        file_size: int,
        file_hash: str,
        network_zone: str,
        scan_result: ScanResult,
        adapter
    ) -> Dict[str, Any]:
        """Procesar archivo grande en múltiples batches."""
        
        import time
        start_time = time.time()
        
        # 1. Crear registro de importación (solo 1 vez)
        import anyio
        logger.info("Step 1: Creating scan_import record...")
        create_result = await anyio.to_thread.run_sync(
            lambda: supabase.rpc_with_token(
                'fn_create_scan_import_record',
                access_token,
                {
                    'p_workspace_id': workspace_id,
                    'p_project_id': project_id,
                    'p_file_name': filename,
                    'p_storage_path': storage_path,
                    'p_file_size': file_size,
                    'p_file_hash': file_hash,
                    'p_scanner': adapter.scanner_name,
                    'p_network_zone': network_zone,
                    'p_scan_start': scan_result.scan_start.isoformat() if scan_result.scan_start else None,
                    'p_scan_end': scan_result.scan_end.isoformat() if scan_result.scan_end else None,
                    'p_total_hosts': scan_result.total_hosts,
                    'p_total_findings': scan_result.total_findings
                }
            )
        )
        
        if not create_result.get('success'):
            raise RPCError('fn_create_scan_import_record', create_result.get('error', 'Unknown error'))
        
        scan_import_id = create_result.get('scan_import_id')
        logger.info(f"Created scan_import: {scan_import_id}")
        
        try:
            # 2. Procesar assets (todos juntos, usualmente son pocos)
            assets_json = self._serialize_assets(scan_result.assets)
            
            # 2.5. Obtener mapeo plugin_id -> vulnerability_id
            plugin_to_vuln_id = {}
            try:
                # Extraer vulnerabilidades únicas
                findings_json_all = self._serialize_findings(scan_result.findings)
                unique_vulns = self._extract_unique_vulnerabilities(findings_json_all)
                
                if unique_vulns:
                    logger.info(f"Processing {len(unique_vulns)} unique vulnerabilities for catalog...")
                    
                    # Importar translation_service
                    from app.services.translation_service import translation_service
                    
                    # Proceso completo: consultar existentes, traducir nuevas, insertar
                    result = await translation_service.translate_new_vulnerabilities(
                        access_token=access_token,
                        scanner=adapter.scanner_name,
                        vulnerabilities=unique_vulns
                    )
                    
                    plugin_to_vuln_id = result['plugin_to_vuln_id']
                    catalog_stats = result['stats']
                    logger.info(f"Catalog stats: {catalog_stats}")
                    
            except Exception as e:
                logger.warning(f"Translation step failed (continuing without vulnerability_id): {e}")
            
            # 3. Procesar findings en batches
            findings = scan_result.findings
            total_findings = len(findings)
            total_created = 0
            total_updated = 0
            total_reopened = 0
            batch_number = 0
            
            for i in range(0, total_findings, self.BATCH_SIZE):
                batch_number += 1
                batch_findings = findings[i:i + self.BATCH_SIZE]
                findings_json = self._serialize_findings(batch_findings)
                
                # Asignar vulnerability_id a los findings del batch
                findings_json = self._assign_vulnerability_ids(findings_json, plugin_to_vuln_id)
                
                # Solo enviar assets en el primer batch
                batch_assets = assets_json if i == 0 else []
                
                # Log temporal para debug
                if i == 0:
                    logger.info(f"Sending {len(batch_assets)} assets in first batch")
                    if batch_assets:
                        logger.info(f"Sample asset: {batch_assets[0] if batch_assets else 'None'}")
                
                logger.info(f"Processing batch {batch_number}: {len(batch_findings)} findings...")
                
                # Usar conexión directa a PostgreSQL para evitar timeout de PostgREST (10s)
                postgres_client = get_postgres_client()
                batch_result = await postgres_client.execute_function(
                    'fn_process_scan_batch',
                    {
                        'p_scan_import_id': scan_import_id,
                        'p_workspace_id': workspace_id,
                        'p_project_id': project_id,
                        'p_assets': batch_assets,
                        'p_findings': findings_json,
                        'p_batch_number': batch_number
                    }
                )
                
                if not batch_result.get('success'):
                    raise RPCError('fn_process_scan_batch', batch_result.get('error', 'Batch failed'))
                
                # Acumular estadísticas
                batch_stats = batch_result.get('batch_stats', {})
                total_created += batch_stats.get('findings_created', 0)
                total_updated += batch_stats.get('findings_updated', 0)
                total_reopened += batch_stats.get('findings_reopened', 0)
                
                logger.info(
                    f"Batch {batch_number} completed in {batch_result.get('processing_time_ms', 0)}ms. "
                    f"Created: {batch_stats.get('findings_created', 0)}, "
                    f"Updated: {batch_stats.get('findings_updated', 0)}"
                )
            
            # 4. Finalizar importación
            logger.info("Finalizing scan import...")
            finalize_result = await anyio.to_thread.run_sync(
                lambda: supabase.rpc_with_token(
                    'fn_finalize_scan_import',
                    access_token,
                    {
                        'p_scan_import_id': scan_import_id,
                        'p_project_id': project_id,
                        'p_total_findings_created': total_created,
                        'p_total_findings_updated': total_updated,
                        'p_total_assets': len(scan_result.assets)
                    }
                )
            )
            
            if not finalize_result.get('success'):
                raise RPCError('fn_finalize_scan_import', finalize_result.get('error', 'Finalize failed'))
            
            total_time_ms = int((time.time() - start_time) * 1000)
            
            logger.info(
                f"Batch import completed in {total_time_ms}ms. "
                f"Batches: {batch_number}, Created: {total_created}, Updated: {total_updated}"
            )
            
            return {
                "scan_import_id": scan_import_id,
                "scanner": adapter.scanner_name,
                "status": "processed",
                "processing_time_ms": total_time_ms,
                "mode": "batch",
                "batches_processed": batch_number,
                "batch_size": self.BATCH_SIZE,
                "assets_upserted": len(scan_result.assets),
                "findings_total": total_findings,
                "findings_created": total_created,
                "findings_updated": total_updated,
                "findings_reopened": total_reopened,
                "findings_by_severity": finalize_result.get('findings_by_severity', {}),
                "scan_info": {
                    "name": scan_result.scan_name,
                    "policy": scan_result.scan_policy,
                    "start": scan_result.scan_start.isoformat() if scan_result.scan_start else None,
                    "end": scan_result.scan_end.isoformat() if scan_result.scan_end else None,
                },
                "warnings": scan_result.warnings,
                "errors": scan_result.errors
            }
            
        except Exception as e:
            # Marcar como fallido
            logger.error(f"Batch processing failed: {e}")
            try:
                await supabase.rpc_with_token(
                    'fn_fail_scan_import',
                    access_token,
                    {
                        'p_scan_import_id': scan_import_id,
                        'p_error_message': str(e)
                    }
                )
            except:
                pass
            raise
    
    def _serialize_assets(self, assets: List) -> List[Dict[str, Any]]:
        """
        Serialize RawAsset objects to JSON-compatible dicts.
        Includes ALL fields that the adapter extracts.
        """
        result = []
        for asset in assets:
            result.append({
                # Identification
                'identifier': asset.identifier,
                'name': asset.name or asset.hostname or str(asset.ip_address) if asset.ip_address else asset.identifier,
                'hostname': asset.hostname,
                'fqdn': getattr(asset, 'fqdn', None),
                'netbios_name': getattr(asset, 'netbios_name', None),
                'mac_address': getattr(asset, 'mac_address', None),
                'ip_address': str(asset.ip_address) if asset.ip_address else None,
                
                # Classification
                'asset_type': asset.asset_type.value if hasattr(asset.asset_type, 'value') else str(asset.asset_type),
                'os_name': asset.os_name,
                'os_family': getattr(asset, 'os_family', None),
                
                # Scan info
                'scan_start': asset.scan_start.isoformat() if asset.scan_start else None,
                'scan_end': asset.scan_end.isoformat() if asset.scan_end else None,
                
                # Metadata (everything else from Nessus)
                'metadata': getattr(asset, 'metadata', {}) or {}
            })
        return result
    
    def _serialize_findings(self, findings: List) -> List[Dict[str, Any]]:
        """
        Serialize RawFinding objects to JSON-compatible dicts.
        Includes ALL fields that the adapter extracts.
        """
        result = []
        for f in findings:
            # Get severity as string
            severity = f.severity.value if hasattr(f.severity, 'value') else str(f.severity)
            
            # Build extras dict with all exploit info
            extras = getattr(f, 'extras', {}) or {}
            
            result.append({
                # Core identification
                'fingerprint': f.fingerprint,
                'asset_identifier': f.asset_identifier,
                'scanner': f.scanner,
                'scanner_finding_id': f.scanner_finding_id or f.plugin_id,
                
                # Content
                'title': f.title,
                'synopsis': getattr(f, 'synopsis', None),
                'description': f.description,
                'solution': f.solution,
                
                # Severity
                'severity': severity,
                'original_severity': f.original_severity,
                'risk_factor': getattr(f, 'risk_factor', None),
                
                # Location
                'hostname': f.asset_identifier if not self._is_ip(f.asset_identifier) else None,
                'ip_address': f.asset_identifier if self._is_ip(f.asset_identifier) else None,
                'port': f.port,
                'protocol': f.protocol,
                'service': f.service,
                'location': f.location,
                
                # Classification
                'cwe': f.cwe,
                'cves': f.cves if f.cves else None,
                
                # CVSS v2
                'cvss_score': f.cvss_score,
                'cvss_vector': f.cvss_vector,
                
                # CVSS v3 (IMPORTANTE - antes no se guardaba)
                'cvss3_score': getattr(f, 'cvss3_score', None),
                'cvss3_vector': getattr(f, 'cvss3_vector', None),
                
                # References (IMPORTANTE - antes no se guardaba)
                'references': getattr(f, 'references', None),
                'reference_ids': getattr(f, 'reference_ids', None),
                
                # Plugin info
                'plugin_id': f.plugin_id,
                'plugin_name': getattr(f, 'plugin_name', None),
                'plugin_family': getattr(f, 'plugin_family', None),
                'plugin_type': getattr(f, 'plugin_type', None),
                'plugin_output': getattr(f, 'plugin_output', None),
                
                # Raw output (truncated)
                'raw_output': (f.plugin_output[:500] if f.plugin_output else None),
                
                # Extras con info de exploits y fechas
                'extras': extras
            })
        return result
    
    def _is_ip(self, value: str) -> bool:
        """Check if value looks like an IP address."""
        if not value:
            return False
        parts = value.split('.')
        if len(parts) == 4:
            try:
                return all(0 <= int(p) <= 255 for p in parts)
            except ValueError:
                pass
        return False
    
    async def _check_duplicate(
        self,
        access_token: str,
        workspace_id: str,
        file_hash: str,
        project_id: Optional[str] = None
    ) -> bool:
        """Check if file hash already exists in the same project."""
        try:
            import anyio
            
            query = supabase.service.table('scan_imports').select('id').eq(
                'workspace_id', workspace_id
            ).eq('file_hash', file_hash)
            
            # Si se proporciona project_id, verificar duplicados solo en ese proyecto
            if project_id:
                query = query.eq('project_id', project_id)
            
            result = await anyio.to_thread.run_sync(lambda: query.execute())
            
            return len(result.data) > 0
        except Exception as e:
            logger.error(f"Error checking duplicate: {e}")
            return False
    
    async def _upload_to_storage(
        self,
        workspace_id: str,
        file_content: bytes,
        filename: str
    ) -> str:
        """Upload scan file to Supabase Storage."""
        try:
            timestamp = datetime.utcnow().strftime("%Y%m%d_%H%M%S")
            unique_id = str(uuid4())[:8]
            storage_path = f"{workspace_id}/scans/{timestamp}_{unique_id}_{filename}"
            
            supabase.service.storage.from_(settings.STORAGE_BUCKET).upload(
                storage_path,
                file_content,
                {"content-type": "application/octet-stream"}
            )
            
            return storage_path
            
        except Exception as e:
            logger.error(f"Storage upload error: {e}")
            raise StorageError(f"Failed to upload file: {str(e)}", "upload")
    
    async def list_scans(
        self,
        access_token: str,
        project_id: str,
        page: int = 1,
        per_page: int = 20
    ) -> Dict[str, Any]:
        """List scans for a project."""
        try:
            import anyio
            result = await anyio.to_thread.run_sync(
                lambda: supabase.rpc_with_token(
                    'fn_list_scans',
                    access_token,
                    {
                        'p_project_id': project_id,
                        'p_page': page,
                        'p_per_page': per_page
                    }
                )
            )
            return result
        except Exception as e:
            logger.error(f"Error listing scans: {e}")
            raise RPCError('fn_list_scans', str(e))
    
    async def get_scan_diff(
        self,
        access_token: str,
        scan_id: str
    ) -> Dict[str, Any]:
        """Get diff between scan and previous scan."""
        try:
            import anyio
            result = await anyio.to_thread.run_sync(
                lambda: supabase.rpc_with_token(
                    'fn_get_scan_diff',
                    access_token,
                    {'p_scan_id': scan_id}
                )
            )
            return result
        except Exception as e:
            logger.error(f"Error getting scan diff: {e}")
            raise RPCError('fn_get_scan_diff', str(e))
    
    async def get_scan_diff_summary(
        self,
        access_token: str,
        scan_id: str
    ) -> Dict[str, Any]:
        """Get scan diff summary only (lazy)."""
        try:
            import anyio
            result = await anyio.to_thread.run_sync(
                lambda: supabase.rpc_with_token(
                    'fn_get_scan_diff_summary',
                    access_token,
                    {'p_scan_id': scan_id}
                )
            )
            return result
        except Exception as e:
            logger.error(f"Error getting scan diff summary: {e}")
            raise RPCError('fn_get_scan_diff_summary', str(e))

    async def get_scan_diff_findings(
        self,
        access_token: str,
        scan_id: str,
        diff_type: str,
        page: int = 1,
        per_page: int = 20
    ) -> Dict[str, Any]:
        """Get paginated findings for specific diff type."""
        try:
            import anyio
            result = await anyio.to_thread.run_sync(
                lambda: supabase.rpc_with_token(
                    'fn_get_scan_diff_findings',
                    access_token,
                    {
                        'p_scan_id': scan_id,
                        'p_diff_type': diff_type,
                        'p_page': page,
                        'p_per_page': per_page
                    }
                )
            )
            return result
        except Exception as e:
            logger.error(f"Error getting diff findings: {e}")
            raise RPCError('fn_get_scan_diff_findings', str(e))
    
    async def generate_excel_report(
        self,
        access_token: str,
        project_id: str,
        include_info: bool = False,
        include_evidence: bool = False
    ) -> bytes:
        """
        Generate Excel report for project findings.
        
        Now includes CVSS v3 and more fields.
        """
        # Get all findings
        import anyio
        findings_result = await anyio.to_thread.run_sync(
            lambda: supabase.rpc_with_token(
                'fn_list_findings',
                access_token,
                {
                    'p_project_id': project_id,
                    'p_page': 1,
                    'p_per_page': 10000
                }
            )
        )
        
        findings = findings_result.get('data', [])
        
        if not include_info:
            findings = [f for f in findings if f.get('severity') != 'Info']
        
        # Sort by severity
        severity_order = {"Critical": 0, "High": 1, "Medium": 2, "Low": 3, "Info": 4}
        findings.sort(key=lambda f: severity_order.get(f.get('severity', 'Info'), 5))
        
        # Create workbook
        wb = Workbook()
        ws = wb.active
        ws.title = "Vulnerabilidades"
        
        # Headers (expanded)
        headers = [
            "Folio", "Vulnerabilidad", "Sinopsis", "Descripción", "Severidad",
            "CVSS v3", "CVSS v2", "Dirección IP", "Puerto", "Protocolo",
            "Hostname", "Servicio", "Recomendación", "CVEs", "CWE",
            "Exploit Disponible", "Plugin Family", "Estado", "Primera Detección",
            "Última Detección", "Referencias"
        ]
        
        header_font = Font(bold=True, color="FFFFFF")
        header_fill = PatternFill(start_color="333333", end_color="333333", fill_type="solid")
        
        for col, header in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col, value=header)
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
        
        # Data rows
        for row_num, finding in enumerate(findings, 2):
            col = 1
            ws.cell(row=row_num, column=col, value=finding.get('folio', '')); col += 1
            ws.cell(row=row_num, column=col, value=finding.get('title', '')); col += 1
            ws.cell(row=row_num, column=col, value=finding.get('synopsis', '')[:200] if finding.get('synopsis') else ''); col += 1
            ws.cell(row=row_num, column=col, value=finding.get('description', '')[:500] if finding.get('description') else ''); col += 1
            
            # Severity with color
            severity = finding.get('severity', 'Info')
            severity_cell = ws.cell(row=row_num, column=col, value=severity); col += 1
            if severity in SEVERITY_COLORS:
                severity_cell.fill = PatternFill(
                    start_color=SEVERITY_COLORS[severity],
                    end_color=SEVERITY_COLORS[severity],
                    fill_type="solid"
                )
                if severity in ["Critical", "High"]:
                    severity_cell.font = Font(color="FFFFFF", bold=True)
            
            # CVSS scores
            cvss3 = finding.get('cvss3_score')
            cvss2 = finding.get('cvss_score')
            ws.cell(row=row_num, column=col, value=cvss3 if cvss3 else ''); col += 1
            ws.cell(row=row_num, column=col, value=cvss2 if cvss2 else ''); col += 1
            
            # Network info
            ws.cell(row=row_num, column=col, value=finding.get('ip_address', '')); col += 1
            ws.cell(row=row_num, column=col, value=finding.get('port', '')); col += 1
            ws.cell(row=row_num, column=col, value=finding.get('protocol', '')); col += 1
            ws.cell(row=row_num, column=col, value=finding.get('hostname', '')); col += 1
            ws.cell(row=row_num, column=col, value=finding.get('service', '')); col += 1
            
            # Solution
            ws.cell(row=row_num, column=col, value=finding.get('solution', '')[:500] if finding.get('solution') else ''); col += 1
            
            # CVEs and CWE
            cves = finding.get('cves', [])
            ws.cell(row=row_num, column=col, value=', '.join(cves) if cves else ''); col += 1
            ws.cell(row=row_num, column=col, value=finding.get('cwe', '')); col += 1
            
            # Exploit info
            exploit = 'Sí' if finding.get('exploit_available') else 'No'
            ws.cell(row=row_num, column=col, value=exploit); col += 1
            
            # Plugin family
            ws.cell(row=row_num, column=col, value=finding.get('plugin_family', '')); col += 1
            
            # Status
            ws.cell(row=row_num, column=col, value=finding.get('status', '')); col += 1
            
            # Dates
            first_seen = finding.get('first_seen', '')
            last_seen = finding.get('last_seen', '')
            ws.cell(row=row_num, column=col, value=first_seen[:10] if first_seen else ''); col += 1
            ws.cell(row=row_num, column=col, value=last_seen[:10] if last_seen else ''); col += 1
            
            # References
            refs = finding.get('references', [])
            ws.cell(row=row_num, column=col, value='\n'.join(refs[:3]) if refs else ''); col += 1
        
        # Column widths
        widths = [10, 45, 40, 50, 10, 8, 8, 15, 8, 8, 20, 12, 45, 30, 12, 10, 20, 12, 12, 12, 40]
        for col, width in enumerate(widths, 1):
            if col <= 26:  # A-Z
                ws.column_dimensions[chr(64 + col)].width = width
        
        # Freeze header row
        ws.freeze_panes = 'A2'
        
        # Auto-filter
        ws.auto_filter.ref = ws.dimensions
        
        # Save to bytes
        output = io.BytesIO()
        wb.save(output)
        output.seek(0)
        
        return output.getvalue()
    
    async def generate_executive_summary(
        self,
        access_token: str,
        project_id: str
    ) -> Dict[str, Any]:
        """Generate executive summary stats for a project."""
        try:
            import anyio
            result = await anyio.to_thread.run_sync(
                lambda: supabase.rpc_with_token(
                    'fn_get_dashboard_project',
                    access_token,
                    {'p_project_id': project_id}
                )
            )
            return result
        except Exception as e:
            logger.error(f"Error generating summary: {e}")
            raise RPCError('fn_get_dashboard_project', str(e))


# Singleton instance
import_service = ImportService()
